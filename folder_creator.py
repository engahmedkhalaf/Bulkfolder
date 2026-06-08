import os
import csv
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Optional Excel support — degrades gracefully if openpyxl isn't installed.
try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None
    load_workbook = None


# ----------------------------------------------------------------------
# Core actions
# ----------------------------------------------------------------------
def browse_directory():
    directory = filedialog.askdirectory()
    if directory:
        base_dir_var.set(directory)


HEADER_WORDS = (
    "path", "paths", "folder", "folders", "name", "directory",
    "level", "level 1", "level1", "subfolder", "sub-folder", "parent",
)


def _read_rows_from_csv(filepath):
    """Read all rows (lists of cell strings) from a CSV file."""
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            rows.append([(c.strip() if c is not None else "") for c in row])
    return rows


def _read_rows_from_excel(filepath):
    """Read all rows (lists of cell strings) from the active Excel sheet."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([(str(c).strip() if c is not None else "") for c in row])
    wb.close()
    return rows


def _rows_to_paths(rows):
    """Turn spreadsheet rows into folder paths.

    Two layouts are supported automatically:

    * Single column  -> each cell is a full path (slashes allowed),
                        e.g.  01-Native/Documents/Invoices
    * Multiple columns -> each column is one level deeper in the tree.
                        Blank cells inherit the value above them (fill-down),
                        so you only type a parent once.

        01-Native | Documents | Invoices
                  |           | Receipts      -> 01-Native/Documents/Receipts
                  | Contracts                 -> 01-Native/Contracts
        02-Exchange | Imports                 -> 02-Exchange/Imports
    """
    paths = []
    inherited = []  # remembered value for each column level (for fill-down)

    for raw in rows:
        # Drop trailing empty cells so we know the real depth of this row.
        cells = list(raw)
        while cells and cells[-1] == "":
            cells.pop()
        if not cells:
            continue

        # Single-column file: treat the cell as a complete path.
        if len(cells) == 1:
            paths.append(cells[0])
            inherited = []
            continue

        # Multi-column: build the path level by level with fill-down.
        depth = len(cells)
        if len(inherited) < depth:
            inherited += [""] * (depth - len(inherited))

        for i in range(depth):
            if cells[i]:
                inherited[i] = cells[i]
                # A new value at this level invalidates everything deeper.
                for j in range(i + 1, len(inherited)):
                    inherited[j] = ""

        parts = [inherited[i] for i in range(depth)]
        if all(parts):  # every level filled -> valid path
            paths.append("/".join(parts))

    return paths


def _strip_header(rows):
    """Remove a leading header row if the first cell looks like a label."""
    if rows and rows[0] and rows[0][0].lower() in HEADER_WORDS:
        return rows[1:]
    return rows


def import_file():
    """Import a list of folder paths from a CSV or Excel file."""
    filetypes = [("CSV files", "*.csv")]
    if EXCEL_AVAILABLE:
        filetypes.insert(0, ("Excel files", "*.xlsx *.xlsm"))
    filetypes.append(("All files", "*.*"))

    filepath = filedialog.askopenfilename(
        title="Import folder structure", filetypes=filetypes
    )
    if not filepath:
        return

    try:
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            if not EXCEL_AVAILABLE:
                messagebox.showerror(
                    "Excel not supported",
                    "Excel import needs the 'openpyxl' package.\n\n"
                    "Install it with:  pip install openpyxl",
                )
                return
            rows = _read_rows_from_excel(filepath)
        else:
            rows = _read_rows_from_csv(filepath)
    except Exception as e:
        messagebox.showerror("Import failed", f"Could not read the file:\n{e}")
        return

    rows = _strip_header(rows)
    paths = _rows_to_paths(rows)

    if not paths:
        messagebox.showwarning("Nothing found", "No folder paths were found in that file.")
        return

    existing = folder_text.get("1.0", tk.END).strip()
    new_block = "\n".join(paths)
    if existing:
        folder_text.insert(tk.END, "\n" + new_block)
    else:
        folder_text.insert("1.0", new_block)

    set_status(f"Imported {len(paths)} path(s) from {os.path.basename(filepath)}")


def export_file():
    """Export the current folder list to CSV or Excel."""
    folder_data = folder_text.get("1.0", tk.END).strip()
    if not folder_data:
        messagebox.showwarning("Nothing to export", "The folder list is empty.")
        return

    paths = [ln.strip() for ln in folder_data.splitlines() if ln.strip()]

    filetypes = [("CSV files", "*.csv")]
    default_ext = ".csv"
    if EXCEL_AVAILABLE:
        filetypes.insert(0, ("Excel files", "*.xlsx"))
        default_ext = ".xlsx"

    filepath = filedialog.asksaveasfilename(
        title="Export folder structure",
        defaultextension=default_ext,
        filetypes=filetypes,
        initialfile=f"folder_structure_{datetime.date.today():%Y%m%d}",
    )
    if not filepath:
        return

    try:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".xlsx":
            if not EXCEL_AVAILABLE:
                messagebox.showerror(
                    "Excel not supported",
                    "Excel export needs 'openpyxl'.\n\npip install openpyxl",
                )
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Folders"
            ws.append(["Path"])
            for p in paths:
                ws.append([p])
            ws.column_dimensions["A"].width = 50
            wb.save(filepath)
        else:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Path"])
                for p in paths:
                    writer.writerow([p])
    except Exception as e:
        messagebox.showerror("Export failed", f"Could not write the file:\n{e}")
        return

    set_status(f"Exported {len(paths)} path(s) to {os.path.basename(filepath)}")


def clear_text():
    if folder_text.get("1.0", tk.END).strip():
        if messagebox.askyesno("Clear", "Clear the folder list?"):
            folder_text.delete("1.0", tk.END)
            set_status("Cleared.")


def generate_folders():
    base_dir = base_dir_var.get().strip()
    if not base_dir:
        messagebox.showerror("Error", "Please select a Base Directory first.")
        return
    if not os.path.exists(base_dir):
        messagebox.showerror("Error", "The selected Base Directory does not exist.")
        return

    folder_data = folder_text.get("1.0", tk.END).strip()
    if not folder_data:
        messagebox.showerror("Error", "The Folder Structure text box is empty.")
        return

    lines = folder_data.splitlines()
    success_count = 0
    errors = []

    for line in lines:
        path_suffix = line.strip()
        if not path_suffix:
            continue
        full_path = os.path.normpath(os.path.join(base_dir, path_suffix))
        try:
            os.makedirs(full_path, exist_ok=True)
            success_count += 1
        except Exception as e:
            errors.append(f"'{path_suffix}': {str(e)}")

    if not errors:
        set_status(f"Created {success_count} folder(s) successfully.")
        messagebox.showinfo("Success", f"Successfully created all {success_count} folder structure(s)!")
    else:
        set_status(f"Done with {len(errors)} error(s).")
        error_msg = "\n".join(errors)
        messagebox.showwarning(
            "Completed with Errors",
            f"Successfully created: {success_count}\nFailed: {len(errors)}\n\nDetails:\n{error_msg}",
        )


def set_status(msg):
    status_var.set(f"  {msg}")


# ----------------------------------------------------------------------
# UI Setup
# ----------------------------------------------------------------------
root = tk.Tk()
root.title("Bulk Folder Creator")
root.geometry("620x560")
root.minsize(500, 420)

style = ttk.Style()
style.theme_use("clam")

# A little accent styling for the primary action button.
ACCENT = "#2563eb"
style.configure("Accent.TButton", font=("Helvetica", 10, "bold"),
                foreground="white", background=ACCENT, padding=6)
style.map("Accent.TButton", background=[("active", "#1d4ed8")])

main_frame = ttk.Frame(root, padding="20")
main_frame.pack(fill="both", expand=True)

# 1. Base Directory
lbl_base = ttk.Label(main_frame, text="Base Directory:", font=("Helvetica", 10, "bold"))
lbl_base.pack(anchor="w", pady=(0, 5))

dir_frame = ttk.Frame(main_frame)
dir_frame.pack(fill="x", pady=(0, 15))

base_dir_var = tk.StringVar()
entry_dir = ttk.Entry(dir_frame, textvariable=base_dir_var, font=("Helvetica", 10))
entry_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))

btn_browse = ttk.Button(dir_frame, text="Browse...", command=browse_directory)
btn_browse.pack(side="right")

# 2. Folder Structure header + toolbar (import/export/clear)
header_frame = ttk.Frame(main_frame)
header_frame.pack(fill="x", pady=(0, 2))

lbl_structure = ttk.Label(header_frame, text="Folder Structure:", font=("Helvetica", 10, "bold"))
lbl_structure.pack(side="left", anchor="w")

btn_clear = ttk.Button(header_frame, text="Clear", command=clear_text)
btn_clear.pack(side="right")

btn_export = ttk.Button(header_frame, text="Export \u25bc", command=export_file)
btn_export.pack(side="right", padx=(0, 6))

btn_import = ttk.Button(header_frame, text="Import \u25b2", command=import_file)
btn_import.pack(side="right", padx=(0, 6))

# Hint text
excel_note = "CSV/Excel" if EXCEL_AVAILABLE else "CSV (install openpyxl for Excel)"
lbl_hint = ttk.Label(
    main_frame,
    text=(
        "\u2022 Single: NewFolder\n"
        "\u2022 Subfolders: Client_A/Documents/Invoices\n"
        "\u2022 Multiple: one path per line\n"
        f"\u2022 Import/Export uses {excel_note} (first column = path)."
    ),
    justify="left",
    foreground="#666666",
    font=("Helvetica", 9, "italic"),
)
lbl_hint.pack(anchor="w", pady=(4, 10))

# Scrollable text area
text_frame = ttk.Frame(main_frame)
text_frame.pack(fill="both", expand=True, pady=(0, 15))

scrollbar = ttk.Scrollbar(text_frame)
scrollbar.pack(side="right", fill="y")

folder_text = tk.Text(text_frame, yscrollcommand=scrollbar.set, font=("Consolas", 10), undo=True)
folder_text.pack(side="left", fill="both", expand=True)
scrollbar.config(command=folder_text.yview)

# 3. Create button
btn_create = ttk.Button(main_frame, text="Create Folders", command=generate_folders, style="Accent.TButton")
btn_create.pack(fill="x", ipady=6)

# 4. Status bar
status_var = tk.StringVar(value="  Ready.")
status_bar = ttk.Label(root, textvariable=status_var, relief="sunken", anchor="w",
                       font=("Helvetica", 9), foreground="#444444")
status_bar.pack(side="bottom", fill="x")

root.mainloop()
